import pandas as pd, json, os
from bs4 import BeautifulSoup
import requests
from lxml import etree
import openpyxl
from openpyxl import load_workbook
from tqdm import tqdm


def scrap(url):
    def find(soup, n):
        main_div = soup.find("div", {"class": "container body-content"}).find_all("div", {"class": "hidden-print"})[-2]
        if n >= 0 and n <= 6:
            a = main_div.find_all("div", {"class": "row"})[0].find("div", {"class": "col-xs-12"}).find_all("p")[n].text
            return a
        if n == 7:
            a = main_div.find_all("div", {"class": "row"})[1].find("div", {"id": "groupeaction-item"}).find("div", {"class": "panel panel-default"}).find("div", {"id": "groupeaction-body"}).find("div", {"class": "col-md-12"}).text
            return a
        if n == 8:
            a = main_div.find_all("div", {"class": "row"})[3].find("div", {"id": "sousGroup-item"}).find_all("div", {"class": "panel panel-default"})[0].find("div", {"class": "panel-body"}).text
            return a
        if n == 9:
            a = main_div.find_all("div", {"class": "row"})[3].find("div", {"id": "sousGroup-item"}).find_all("div", {"class": "panel panel-default"})[1].find("div", {"class": "panel-body"}).text
            return a

    info = [[], [], [], [], [], [], [], [], [], []]

    webpage = requests.get(url)
    soup = BeautifulSoup(webpage.content, "html.parser")
    dom = etree.HTML(str(soup))
    search1 = soup.find("a", {"href": "#sousGroup-1-body"})
    search2 = soup.find("a", {"href": "#sousGroup-2-body"})
    l = 8
    if search1 != None:
        l += 1
    if search2 != None:
        l += 1
    for i in range(l):
        a = find(soup, i)
        try:
            a = a.replace('  ', '').replace('\n', '').replace('\r', '').replace('\t', '')
        except Exception as e:
            print(e)
        info[i].append(a)

    df = pd.read_html(webpage.text)

    names = soup.find("div", {"id": "demandeurs-body"}).find("div", {"class": "panel-body"}).find_all("div", {"class": "panel-group item-group"})
    cfta = {}
    for i in range(len(names)):
        name = names[i].find("div", {"class": "panel panel-default"}).find("div", {"class": "panel-heading"}).find("span", {"class": "hidden-xs"}).text
        cfta[i] = name

    cftd = {}
    names = soup.find("div", {"id": "defendeurs-body"}).find("div", {"class": "panel-body"}).find_all("div", {"class": "panel-group item-group"})
    for i in range(len(names)):
        name = names[i].find("div", {"class": "panel panel-default"}).find("div", {"class": "panel-heading"}).find("span", {"class": "hidden-xs"}).text
        cftd[i] = name

    template = ["Name", "E-mail", "Phone", "Adresse", "Name of Law Firm", "Website"]
    d = df[1].to_dict(orient='dict')
    c = df[2].to_dict(orient='dict')
    for t in template:
        if t not in d:
            d[t] = {0: ""}
    for t in template:
        if t not in c:
            c[t] = {0: ""}
    docs = df[0].to_dict(orient='dict')

    rows = soup.find("table", {"id": "tableau-documents"}).find("tbody").find_all("tr", {"role": "row"})
    links = [[], []]
    for row in rows:
        links[1].insert(0, f'https://www.registredesactionscollectives.quebec/{row.find("td").a["href"]}')
    links[1].reverse()
    for doc in docs["Document"]:
        links[0].append(docs["Document"][doc])
    caseID = url.split('NoDossier=')[1]

    if len(info[8]) == 0:
        info[8].append("")
    if len(info[9]) == 0:
        info[9].append("")

    for i in range(len(links[0])):
        if '/' in links[0][i]:
            links[0][i] = links[0][i].replace('/', '()')
        if '"' in links[0][i]:
            links[0][i] = links[0][i].replace('"', '')

    path = f"pdf_files/{caseID}"
    if not os.path.isdir(path):
        os.mkdir(path)

    p = 1
    for i in tqdm(range(len(links[0])), colour="white"):
        path_pdf = f'{path}/{links[0][i]}.pdf'
        if os.path.isfile(path_pdf):
            path_pdf = f'{path}/{links[0][i]}{p}.pdf'
        while os.path.isfile(path_pdf):
            p += 1
            path_pdf = f'{path}/{links[0][i]}{p}.pdf'
        pdf_response = requests.get(links[1][i])
        with open(path_pdf, 'wb') as f:
            f.write(pdf_response.content)

    data = {
        "Application stage": {0: info[0][0]},
        "Courthouse": {0: info[1][0]},
        "Date of filing": {0: info[2][0]},
        "File no.": {0: info[3][0]},
        "Subject": {0: info[4][0]},
        "This case concerns": {0: info[5][0]},
        "Name of parties": {0: info[6][0]},
        "Description of the class instituting the class action": {0: info[7][0]},
        "Documents and proceedings": docs["Document"],
        "Date of document": docs["Date of document"],

        "Counsels for the applicant": cfta,
        "Name ": d["Name"],
        "E-mail ": d["E-mail"],
        "Phone ": d["Phone"],
        "Adresse ": d["Adresse"],
        "Name of Law Firm ": d["Name of Law Firm"],
        "Website ": d["Website"],

        "Counsels for the defence": cftd,
        "Name": c["Name"],
        "E-mail": c["E-mail"],
        "Phone": c["Phone"],
        "Adresse": c["Adresse"],
        "Name of Law Firm": c["Name of Law Firm"],
        "Website": c["Website"],

        "Description of the sub-class 1": {0: info[8][0]}, # TODO: info[8][0] is null if sub-class not in page. Need to past empty symbol instead
        "Description of the sub-class 2": {0: info[9][0]},
    }
    del info

    exist = True
    if not os.path.isfile("rtable.xlsx"):
        wb = openpyxl.Workbook()
        wb.save("rtable.xlsx")
        exist = False

    with open('data.json', 'w+') as f:
        f.write(json.dumps(data))
    with open('data.json', 'r+') as f:
        a = json.loads(f.readline())
    df = pd.DataFrame(a)
    if exist:
        book = load_workbook('rtable.xlsx')
        writer = pd.ExcelWriter('rtable.xlsx', engine='openpyxl')
        writer.book = book
        writer.sheets = {ws.title: ws for ws in book.worksheets}
        for sheetname in writer.sheets:
            df.to_excel(writer, sheet_name=sheetname, startrow=writer.sheets[sheetname].max_row, index=False, header=False)
        writer.save()
    else:
        df.to_excel("rtable.xlsx", index=False)
    os.remove("data.json")
    return True